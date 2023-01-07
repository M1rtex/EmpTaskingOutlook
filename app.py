import sys
import os
import platform
import subprocess
import webbrowser
from configparser import ConfigParser
from datetime import datetime
import time
from PySide2.QtCore import Qt, QTimer, QThread, Signal, Slot, QObject, QDate
from PySide2.QtWidgets import *
import msal
import json

import openpyxl

import graph_api
from components import select_task, select_project, select_staff, senior_app, log_in, icons

GRAPH_API_ENDPOINT = 'https://graph.microsoft.com/v1.0'
TOKEN = ''
# Read cfg file
configparser = ConfigParser()
configparser.read('data/config.cfg')
APP_ID = configparser['azure']['app_id']
GRAPH_SCOPES = [scope for scope in configparser['azure']['scopes'].split(' ') if scope != '']


def open_folder(path: str) -> None:
    if platform.system() == "Windows":
        if os.path.isfile(path):
            subprocess.Popen(["explorer", "/select,", path])
        else:
            os.startfile(path)
    elif platform.system() == "Darwin":
        subprocess.Popen(["open", path])
    else:
        subprocess.Popen(["xdg-open", path])


def check_paterns(values: list, paterns: dict, name: str = '') -> bool:
    for val in values:
        for patern in paterns.keys():
            if val.lower().find(patern) != -1:
                paterns[patern] = True
    if not all(paterns.values()):
        print('--------------------------------')
        print(f'Paterns {name + " " if name else ""}errors:')
        for patern in paterns.keys():
            print(f'{patern}: {paterns[patern]}')
        print('--------------------------------')
    return all(paterns.values())


class LogInApp(QMainWindow, log_in.Ui_MainWindow):
    def __init__(self, main_app):
        super().__init__()
        self.setupUi(self)
        self.setWindowTitle("Авторизация")
        self.setWindowFlags(Qt.WindowStaysOnTopHint)
        self.clipboard = QApplication.clipboard()
        self.main_app = main_app
        self.timer = QTimer()
        self.workthread = GraphWorker()
        self.workthread.signals.on_change_code.connect(self.change_code)
        self.workthread.signals.on_token_received.connect(self.set_token)
        self.workthread.signals.on_token_updated.connect(self.update_token)
        self.workthread.finished.connect(lambda: self.create_timer_task())
        self.workthread.start()
        self.copy_login_code_btn.clicked.connect(self.copy_to_clipboard)

    def copy_to_clipboard(self):
        self.clipboard.setText(self.login_code_lable.text())

    def create_timer_task(self):
        if not self.timer.isActive():
            self.timer = QTimer()
            self.timer.timeout.connect(self.start_token_update)
            self.timer.setTimerType(Qt.VeryCoarseTimer)
            self.timer.start(3_000_000)

    def start_token_update(self):
        self.workthread.start(priority=QThread.LowPriority)

    @Slot(str)
    def change_code(self, code: str):
        self.login_code_lable.setText(code)
        self.show()
        self.main_app.hide()

    @Slot(dict)
    def set_token(self, token: dict):
        global TOKEN
        TOKEN = token
        self.close()
        self.main_app.show()

    @Slot(dict)
    def update_token(self, token: dict):
        global TOKEN
        TOKEN = token
        print("Token updated")


class SelectStaff(QMainWindow, select_staff.Ui_MainWindow):
    def __init__(self, main: object):
        super().__init__()
        self.main = main
        self.setupUi(self)
        self.setWindowTitle("Выбор сотрудника")
        self.submit_btn.clicked.connect(self.submit)

    def show(self) -> None:
        super().show()
        self.staff_listWidget.clear()
        self.staff_listWidget.addItems(self.list_staff())

    def list_staff(self) -> list:
        wb = self.main.open_database(read_only=True)
        ws = wb['Staff']
        result = []
        item = 2
        if not self.main.selected_project:
            while ws[f'B{item}'].value and ws[f'C{item}'].value:
                result.append(f'{ws[f"B{item}"].value} {ws[f"C{item}"].value}')
                item += 1
        else:
            while ws[f'D{item}'].value:
                projects = [i.strip() for i in ws[f'E{item}'].value.split(';') if i.strip() != '']
                if self.main.selected_project in projects:
                    result.append(f'{ws[f"B{item}"].value} {ws[f"C{item}"].value}')
                item += 1
        wb.close()
        return result

    def submit(self) -> None:
        selected_staff = self.staff_listWidget.selectedItems()
        if selected_staff:
            self.main.set_selected_staff(selected_staff[0].text())
            self.close()


class SelectProject(QMainWindow, select_project.Ui_MainWindow):
    def __init__(self, main: object):
        super().__init__()
        self.main = main
        self.setupUi(self)
        self.setWindowTitle("Выбор проекта")
        self.submit_btn.clicked.connect(self.submit)

    def show(self) -> None:
        super().show()
        self.projects_listWidget.clear()
        self.projects_listWidget.addItems(self.list_projects())

    def list_projects(self) -> list:
        wb = self.main.open_database(read_only=True)
        ws = wb['Projects']
        result = []
        item = 2
        while ws[f'A{item}'].value:
            result.append(ws[f'A{item}'].value)
            item += 1
        wb.close()
        return result

    def submit(self) -> None:
        selected_projects = self.projects_listWidget.selectedItems()
        if self.main.selected_staff != {}:
            selected_project = selected_projects[0].text()
            wb = self.main.open_database(read_only=True)
            ws = wb['Staff']
            item = 2
            while ws[f'D{item}'].value:
                if ws[f'A{item}'].value == self.main.selected_staff['Name'] and ws[f'B{item}'].value == \
                        self.main.selected_staff['Surname']:
                    if selected_project in [i.strip() for i in ws[f'D{item}'].value.split(';') if i.strip() != '']:
                        self.main.set_selected_project(selected_project)
                        wb.close()
                        self.close()
                        return None
                item += 1
            self.main.set_selected_staff('')
            wb.close()
        self.main.set_selected_project(selected_projects[0].text())
        self.close()


class SelectTask(QMainWindow, select_task.Ui_MainWindow):
    def __init__(self, main_app: object):
        super().__init__()
        self.main_app = main_app
        self.setupUi(self)
        self.setWindowTitle("Выбор задачи")
        self.submit_btn.clicked.connect(self.submit)

    def show(self) -> None:
        super().show()
        self.tasks_listWidget.clear()
        self.tasks_listWidget.addItems(self.list_tasks())

    def list_tasks(self) -> list:
        wb = self.main_app.open_database(read_only=True)
        ws = wb['Tasks']
        result = []
        item = 2
        if self.main_app.selected_staff:
            ws_staff = wb['Staff']
            tab_num = None
            while ws_staff[f'A{item}'].value:
                if ws_staff[f'B{item}'].value == self.main_app.selected_staff['Name'] and \
                        ws_staff[f'C{item}'].value == self.main_app.selected_staff['Surname']:
                    tab_num = ws_staff[f'A{item}'].value
                    break
                item += 1
            item = 2
            while ws[f'A{item}'].value:
                if ws[f'E{item}'].value == tab_num:
                    result.append(f'{ws[f"B{item}"].value} ({ws[f"A{item}"].value})')
                item += 1
        else:
            while ws[f'A{item}'].value:
                result.append(f'{ws[f"B{item}"].value} ({ws[f"A{item}"].value})')
                item += 1
            wb.close()
        return result

    def submit(self) -> None:
        selected_task = self.tasks_listWidget.selectedItems()[0].text()
        if selected_task:
            self.main_app.set_selected_task(selected_task)
            self.close()


class GraphSignals(QObject):
    on_change_code = Signal(str)
    on_token_received = Signal(dict)
    on_token_updated = Signal(dict)


class GraphWorker(QThread):
    signals = GraphSignals()

    def run(self):
        access_token_cache = msal.SerializableTokenCache()
        # read the token file
        if os.path.exists('data/ms_token.json'):
            access_token_cache.deserialize(open("data/ms_token.json", "r").read())
            token_detail = json.load(open('data/ms_token.json', ))
            token_detail_key = list(token_detail['AccessToken'].keys())[0]
            token_expiration = datetime.fromtimestamp(int(token_detail['AccessToken'][token_detail_key]['expires_on']))
            if datetime.now() > token_expiration:
                os.remove('data/ms_token.json')
                access_token_cache = msal.SerializableTokenCache()

        client = msal.PublicClientApplication(client_id=APP_ID, token_cache=access_token_cache)
        accounts = client.get_accounts()
        if accounts:
            token_response = client.acquire_token_silent(GRAPH_SCOPES, accounts[0])
            self.signals.on_token_updated.emit(token_response)
            with open('data/ms_token.json', 'w') as _f:
                _f.write(access_token_cache.serialize())
        else:
            # authenticating account as usual
            flow = client.initiate_device_flow(scopes=GRAPH_SCOPES)
            if flow['user_code']:
                self.signals.on_change_code.emit(flow['user_code'])
                print(flow['user_code'])
                webbrowser.open('https://microsoft.com/devicelogin')
                token_response = client.acquire_token_by_device_flow(flow)
                self.signals.on_token_received.emit(token_response)
                with open('data/ms_token.json', 'w') as _f:
                    _f.write(access_token_cache.serialize())
            else:
                # TODO: handle error
                sys.exit(flow)


class SeniorApp(QMainWindow, senior_app.Ui_MainWindow):
    def __init__(self):
        super().__init__()
        self.setupUi(self)
        self.setWindowTitle('Enterprise App')
        self.submit_task_icon.setHidden(True)
        self.submit_datetime_icon.setHidden(True)
        self.selected_task_frame.setHidden(True)
        self.submit_task_btn.clicked.connect(self.send_message)
        self.cancel_datetime.clicked.connect(self.reset_selected_task)
        self.submit_datetime_btn.clicked.connect(self.set_task_deadline)
        # Database
        self.database_path = './data/database.xlsx'
        self.database_edited = False
        self.database_default_path = './data/database.xlsx'
        # Temporary information
        self.selected_project = ''
        self.selected_staff = {}
        self.selected_task = {}
        self.staff_paterns = {
            'табельный номер': False,
            'имя': False,
            'фамилия': False,
            'вид деятельности': False,
            'проекты': False,
            'почта': False
        }
        self.projects_paterns = {
            'название': False,
            'начиная': False,
            'срок': False,
            'задачи': False,
        }
        self.tasks_paterns = {
            'ид': False,
            'название': False,
            'начиная': False,
            'заканчивая': False,
            'исполнитель': False,
        }
        self.show()
        self.load_database()

    def send_message(self):
        if self.selected_staff and self.selected_project:
            QTimer.singleShot(0, lambda: self.submit_task_icon.setText('Обработка'))
            self.submit_task_icon.setHidden(False)
            wb = self.open_database(read_only=True)
            start_time = None
            end_time = None
            tab_num = None
            mail = None
            ws_projects = wb['Projects']
            ws_staff = wb['Staff']
            item = 2
            while ws_projects[f'A{item}'].value:
                if ws_projects[f'A{item}'].value == self.selected_project:
                    start_time = ws_projects[f'B{item}'].value
                    end_time = ws_projects[f'C{item}'].value
                    break
                item += 1
            if start_time and end_time:
                start_timestamp = time.mktime(start_time.timetuple())
                end_timestamp = time.mktime(end_time.timetuple())
            else:
                # TODO: handle error
                raise Exception('Start and End time not found')
            item = 2
            while ws_staff[f'A{item}'].value:
                if ws_staff[f'B{item}'].value == self.selected_staff['Name'] and \
                        ws_staff[f'C{item}'].value == self.selected_staff['Surname']:
                    tab_num = ws_staff[f'A{item}'].value
                    mail = ws_staff[f'F{item}'].value
                    break
                item += 1
            if tab_num and mail:
                subject = f'{tab_num}/{start_timestamp}/{end_timestamp}'
                body_ = {}
                if self.par_S_lineedit.text():
                    body_["S"] = self.par_S_lineedit.text()
                if self.par_M_lineedit.text():
                    body_["M"] = self.par_M_lineedit.text()
                if self.par_A_lineedit.text():
                    body_["A"] = self.par_A_lineedit.text()
                if self.par_R_lineedit.text():
                    body_["R"] = self.par_R_lineedit.text()
                if self.par_T_lineedit.text():
                    body_["T"] = self.par_T_lineedit.text()
                body = json.dumps(body_, indent=4)
                res = graph_api.send_text_mail(TOKEN, subject=subject, body=body, recipient=mail)
                print(res.status_code)
                QTimer.singleShot(0, lambda: self.submit_task_icon.setText("Готово"))
                QTimer.singleShot(2000, lambda: self.submit_task_icon.setHidden(True))
            else:
                # TODO: handle error
                raise Exception("No staff info found")

    def create_database(self):
        wb = openpyxl.Workbook()
        ws_projects = wb.create_sheet('Projects')
        ws_projects['A1'] = 'Название'
        ws_projects['B1'] = 'Начиная (с)'
        ws_projects['C1'] = 'Срок (до)'
        ws_projects['D1'] = 'Задачи (ИД)'
        ws_staff = wb.create_sheet('Staff')
        ws_staff['A1'] = 'Табельный номер'
        ws_staff['B1'] = 'Имя'
        ws_staff['C1'] = 'Фамилия'
        ws_staff['D1'] = 'Вид деятельности'
        ws_staff['E1'] = 'Проекты'
        ws_staff['F1'] = 'Почта'
        ws_tasks = wb.create_sheet('Tasks')
        ws_tasks['A1'] = 'ИД'
        ws_tasks['B1'] = 'Название'
        ws_tasks['C1'] = 'Начиная с'
        ws_tasks['D1'] = 'Заканчивая'
        ws_tasks['E1'] = 'Исполнитель'
        wb = self.save_database(wb)
        wb.close()
        self.statusBar.showMessage("Database created", 5000)

    def save_database(self, wb: openpyxl.Workbook) -> openpyxl.Workbook:
        if self.database_edited:
            wb.save(self.database_default_path)
        wb.save(self.database_path)
        return wb

    def open_database(self, read_only=False) -> openpyxl.Workbook:
        try:
            wb = openpyxl.load_workbook(self.database_path, read_only=read_only)
        except FileNotFoundError:
            if self.database_edited:
                try:
                    wb = openpyxl.load_workbook(self.database_default_path, read_only=read_only)
                    main.statusBar.showMessage("Database not found! Check the location! Loaded backup.", 15000)
                except FileNotFoundError:
                    self._database_not_found()
                    wb = openpyxl.load_workbook(self.database_default_path, read_only=read_only)
            else:
                self._database_not_found()
                wb = openpyxl.load_workbook(self.database_default_path, read_only=read_only)
        return wb

    def _database_not_found(self):
        msgBox = QMessageBox()
        msgBox.setText("База данных не найдена")
        msgBox.setInformativeText("Создать новую?")
        applyButton = msgBox.addButton('Ок', msgBox.YesRole)
        discardButton = msgBox.addButton('Ручная проверка', msgBox.RejectRole)
        msgBox.setDefaultButton(applyButton)
        msgBox.exec()
        if msgBox.clickedButton() == applyButton:
            self.create_database()
            self.load_database()
        elif msgBox.clickedButton() == discardButton:
            open_folder(os.getcwd() + '/data')
            sys.exit(0)

    def load_database(self) -> None:
        wb = self.open_database(read_only=True)
        if 'Projects' in wb.sheetnames and 'Staff' in wb.sheetnames and 'Tasks' in wb.sheetnames:
            row1_staff = [val.value for val in wb['Staff'][1]]
            row1_projects = [val.value for val in wb['Projects'][1]]
            row1_tasks = [val.value for val in wb['Tasks'][1]]
            staff_res = check_paterns(row1_staff, self.staff_paterns, "Staff")
            projects_res = check_paterns(row1_projects, self.projects_paterns, "Projects")
            tasks_res = check_paterns(row1_tasks, self.tasks_paterns, "Tasks")
            if staff_res and projects_res and tasks_res:
                wb.close()
                self.statusBar.showMessage("Database OK", 5000)
                return None
        wb.close()
        msgBox = QMessageBox()
        msgBox.setText("База данных повреждена")
        msgBox.setInformativeText("Необходимо вмешательство!")
        okButton = msgBox.addButton('Ок', msgBox.AcceptRole)
        recreateButton = msgBox.addButton('Создать новую', msgBox.RejectRole)
        msgBox.setDefaultButton(okButton)
        msgBox.exec()
        if msgBox.clickedButton() == okButton:
            open_folder(self.database_path if not self.database_edited else self.database_default_path)
            self.close()
            sys.exit(0)
        elif msgBox.clickedButton() == recreateButton:
            msgBox_rec = QMessageBox()
            msgBox_rec.setText('Вы уверены?')
            msgBox_rec.setInformativeText("Эта операция сотрёт все существующие данные!")
            okButton = msgBox_rec.addButton('Ок', msgBox_rec.AcceptRole)
            discardButton = msgBox_rec.addButton('Отмена', msgBox_rec.RejectRole)
            msgBox_rec.exec()
            if msgBox_rec.clickedButton() == okButton:
                os.remove(self.database_default_path)
                self.database_path = self.database_default_path
                self.database_edited = False
                self.create_database()
                self.close()
                sys.exit(200)
            else:
                self.load_database()

    def set_selected_task(self, task_name: str) -> None:
        self.selected_task = {
            'Name': task_name.split('(')[0].strip(),
            'Number': int(task_name[(task_name.find('(') + 1):task_name.find(')')].strip())
        }
        wb = self.open_database(read_only=True)
        ws = wb['Tasks']
        item = 2
        while ws[f'A{item}'].value:
            if ws[f'B{item}'].value == self.selected_task['Name'] and ws[f'A{item}'].value == self.selected_task['Number']:
                start_date = ws[f'C{item}'].value
                end_date = ws[f'D{item}'].value
                if start_date != '' and end_date != '':
                    start = QDate(start_date.year, start_date.month, start_date.day)
                    end = QDate(end_date.year, end_date.month, end_date.day)
                    self.start_dateEdit.setDate(start)
                    self.end_dateEdit.setDate(end)
                wb.close()
                break
            item += 1
        self.select_task_btn.setHidden(True)
        self.task_num_lable.setText(str(self.selected_task['Number']))
        self.selected_task_frame.setHidden(False)

    def reset_selected_task(self) -> None:
        self.selected_task = {}
        self.select_task_btn.setHidden(False)
        self.selected_task_frame.setHidden(True)

    def set_task_deadline(self) -> None:
        start_time = self.start_dateEdit.date().toPython()
        end_time = self.end_dateEdit.date().toPython()
        wb = self.open_database()
        ws = wb['Tasks']
        item = 2
        while ws[f'A{item}'].value:
            if ws[f'B{item}'].value == self.selected_task['Name'] and ws[f'A{item}'].value == self.selected_task['Number']:
                ws[f'C{item}'] = start_time
                ws[f'D{item}'] = end_time
                self.save_database(wb)
                self.statusBar.showMessage(f'Task {self.selected_task["Number"]} deadline updated.', 3000)
                break
            item += 1

    def set_selected_project(self, project_name: str) -> None:
        if project_name == '':
            self.selected_project = ''
            self.proj_lable.setText('не выбрано')
            return None
        self.selected_project = project_name
        self.proj_lable.setText(project_name)

    def set_selected_staff(self, staff: str) -> None:
        if staff == '':
            self.selected_staff = {}
            self.staff_lable.setText('не выбрано')
            return None
        staff = staff.split(' ')
        self.selected_staff = {
            'Name': staff[0],
            'Surname': staff[1]
        }
        self.staff_lable.setText(self.selected_staff['Name'] + ' ' + self.selected_staff['Surname'])


if __name__ == '__main__':
    app = QApplication()
    main = SeniorApp()
    log_in_app = LogInApp(main)
    select_tsk = SelectTask(main)
    select_prj = SelectProject(main)
    select_stf = SelectStaff(main)
    main.select_task_btn.clicked.connect(select_tsk.show)
    main.select_proj_btn.clicked.connect(select_prj.show)
    main.select_staff_btn.clicked.connect(select_stf.show)
    app.exec_()
